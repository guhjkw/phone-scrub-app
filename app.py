"""Phone Scrub — Streamlit app for cleaning skip-trace Excel files."""
from __future__ import annotations

import io
import re
import time

import requests
import streamlit as st
from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# Regex / constants
# ---------------------------------------------------------------------------

# Type label starts with a letter so we don't accidentally match the next
# area code (e.g. "(661)") as a type when numbers lack a label.
PHONE_RE = re.compile(
    r"\((\d{3})\)\s*(\d{3})-(\d{4})\s*(?:\(([A-Za-z]\w*)\))?"
)

# N11 service codes — invalid as area codes or exchanges
N11 = {"211", "311", "411", "511", "611", "711", "811", "911"}

# ---------------------------------------------------------------------------
# Parsing helpers
# ---------------------------------------------------------------------------


def _parse_cell(value) -> list[tuple[str, str]]:
    """Return [(formatted_number, phone_type), ...] from a raw cell value."""
    if value is None:
        return []
    s = str(value).strip()
    if not s or s.upper() == "N/A":
        return []
    return [
        (f"({m.group(1)}) {m.group(2)}-{m.group(3)}", m.group(4) or "")
        for m in PHONE_RE.finditer(s)
    ]


def _nanp_valid(area: str, exchange: str) -> bool:
    return (
        area[0] not in "01"
        and area not in N11
        and exchange[0] not in "01"
        and exchange not in N11
    )


def _structural_pass(area: str, exchange: str, ptype: str) -> tuple[bool, str]:
    """Return (keep, rejection_reason)."""
    if ptype.lower() == "pager":
        return False, "pager"
    if not _nanp_valid(area, exchange):
        return False, "nanp"
    return True, ""


# ---------------------------------------------------------------------------
# IPQS lookup
# ---------------------------------------------------------------------------


def _ipqs_lookup(digits: str, api_key: str, retries: int = 3) -> dict:
    url = f"https://ipqualityscore.com/api/json/phone/{api_key}/{digits}"
    for attempt in range(retries):
        try:
            resp = requests.get(url, params={"country[]": "US"}, timeout=10)
            if resp.status_code == 200:
                d = resp.json()
                if not d.get("success", False):
                    # API-level failure: bad key, exhausted credits, etc.
                    # CLAUDE.md: never remove on success==false — keep and flag.
                    return {
                        "active": None,
                        "valid": None,
                        "line_type": "",
                        "error": d.get("message", "IPQS success=false"),
                    }
                return {
                    "active": d.get("active"),
                    "valid": d.get("valid"),
                    "line_type": (d.get("line_type") or "").lower(),
                    "do_not_call": bool(d.get("do_not_call", False)),
                    "leaked": bool(d.get("leaked", False)),
                    "error": None,
                }
            if resp.status_code == 429:
                time.sleep(2 ** attempt)
                continue
            return {
                "active": None, "valid": None, "line_type": "",
                "do_not_call": False, "leaked": False,
                "error": f"HTTP {resp.status_code}",
            }
        except requests.Timeout:
            if attempt < retries - 1:
                time.sleep(2 ** attempt)
        except Exception as exc:  # noqa: BLE001
            return {"active": None, "valid": None, "line_type": "",
                    "do_not_call": False, "leaked": False, "error": str(exc)}
    return {"active": None, "valid": None, "line_type": "",
            "do_not_call": False, "leaked": False, "error": "max_retries_exceeded"}


def _should_drop(result: dict, drop_voip: bool, drop_landline: bool) -> tuple[bool, str]:
    """Return (drop, reason). On API error we keep the number."""
    if result["error"]:
        return False, ""
    if result["active"] is False or result["valid"] is False:
        return True, "inactive"
    lt = result["line_type"]
    if drop_voip and lt == "voip":
        return True, "voip"
    if drop_landline and lt == "landline":
        return True, "landline"
    return False, ""


# ---------------------------------------------------------------------------
# Core processing
# ---------------------------------------------------------------------------


def process_workbook(
    file_bytes: bytes,
    api_key: str,
    drop_voip: bool,
    drop_landline: bool,
    progress_bar,
) -> tuple[bytes, dict, int]:
    """
    Clean the workbook and return (output_bytes, stats, unique_ipqs_count).

    Three passes:
      1. Structural pre-filter — free, no API calls.
      2. IPQS validation — one call per unique number (deduped).
      3. Rewrite phone cells in place; build Summary sheet.
    """
    wb = load_workbook(io.BytesIO(file_bytes))
    ws = wb.active

    # Identify phone columns dynamically from the header row
    phone_cols: set[int] = {
        cell.column
        for cell in ws[1]
        if cell.value and "phone" in str(cell.value).lower()
    }
    if not phone_cols:
        raise ValueError(
            "No columns containing 'Phone' in the header were found. "
            "Check that the first row is a header row and at least one "
            "column header includes the word 'Phone'."
        )

    stats: dict[str, int] = {
        "total": 0,
        "nanp": 0,
        "pager": 0,
        "inactive": 0,
        "type_filter": 0,
        "api_error": 0,
        "do_not_call": 0,
        "leaked": 0,
        "kept": 0,
    }

    # ------------------------------------------------------------------
    # Pass 1 — structural filter; collect surviving entries per cell and
    # the set of unique E.164 digit strings for IPQS dedup.
    # ------------------------------------------------------------------
    cell_entries: dict[tuple[int, int], list[tuple[str, str]]] = {}
    unique_numbers: set[str] = set()

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column not in phone_cols:
                continue
            parsed = _parse_cell(cell.value)
            if not parsed:
                continue
            passing: list[tuple[str, str]] = []
            for number, ptype in parsed:
                stats["total"] += 1
                m = PHONE_RE.match(number)
                ok, reason = _structural_pass(m.group(1), m.group(2), ptype)
                if not ok:
                    stats["pager" if reason == "pager" else "nanp"] += 1
                    continue
                passing.append((number, ptype))
                unique_numbers.add(re.sub(r"\D", "", number))
            if passing:
                cell_entries[(cell.row, cell.column)] = passing

    # ------------------------------------------------------------------
    # Pass 2 — IPQS validation (one call per unique number)
    # ------------------------------------------------------------------
    candidates = list(unique_numbers)
    cache: dict[str, dict] = {}
    total = len(candidates)

    for i, digits in enumerate(candidates):
        progress_bar.progress(
            i / total if total else 1.0,
            text=f"Validating {i + 1} of {total} unique numbers…",
        )
        cache[digits] = _ipqs_lookup(digits, api_key)
        if cache[digits]["error"]:
            stats["api_error"] += 1

    progress_bar.progress(1.0, text="Writing cleaned file…")

    # Tally do_not_call and leaked across all successful lookups (report only).
    for result in cache.values():
        if not result["error"]:
            if result["do_not_call"]:
                stats["do_not_call"] += 1
            if result["leaked"]:
                stats["leaked"] += 1

    # Resolve which numbers to drop and tally by reason.
    # Separate set for API errors: keep the number but flag as Unverified.
    drop_map: dict[str, str] = {}
    unverified: set[str] = set()
    for digits, result in cache.items():
        if result["error"]:
            unverified.add(digits)
            continue
        drop, reason = _should_drop(result, drop_voip, drop_landline)
        if drop:
            drop_map[digits] = reason

    for reason in drop_map.values():
        if reason == "inactive":
            stats["inactive"] += 1
        else:
            stats["type_filter"] += 1

    # ------------------------------------------------------------------
    # Pass 3 — rewrite phone cells in place (no DataFrame round-trip)
    # ------------------------------------------------------------------
    for (row_idx, col_idx), entries in cell_entries.items():
        surviving = []
        for number, ptype in entries:
            digits = re.sub(r"\D", "", number)
            if digits in drop_map:
                continue
            if digits in unverified:
                # Keep but label so the client knows validation was skipped
                label = f"{ptype} - Unverified" if ptype else "Unverified"
            else:
                label = ptype
            surviving.append((number, label))
        stats["kept"] += len(surviving)
        cell = ws.cell(row=row_idx, column=col_idx)
        if surviving:
            cell.value = "\n".join(
                f"{number} ({label})" if label else number
                for number, label in surviving
            )
        else:
            cell.value = None

    # ------------------------------------------------------------------
    # Summary sheet
    # ------------------------------------------------------------------
    if "Summary" in wb.sheetnames:
        del wb["Summary"]
    ss = wb.create_sheet("Summary")
    removed = stats["nanp"] + stats["pager"] + stats["inactive"] + stats["type_filter"]
    for row in [
        ["Metric", "Count"],
        ["Total numbers found", stats["total"]],
        ["Removed — NANP invalid", stats["nanp"]],
        ["Removed — Pager", stats["pager"]],
        ["Removed — IPQS inactive / disconnected", stats["inactive"]],
        ["Removed — VOIP / Landline type filter", stats["type_filter"]],
        ["Total removed", removed],
        ["Numbers kept", stats["kept"]],
        ["", ""],
        ["Flagged Do Not Call (kept, not removed)", stats["do_not_call"]],
        ["Flagged leaked credentials (kept, not removed)", stats["leaked"]],
        ["", ""],
        ["Unique numbers sent to IPQS", total],
        ["IPQS API errors (kept as Unverified)", stats["api_error"]],
    ]:
        ss.append(row)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue(), stats, total


# ---------------------------------------------------------------------------
# Streamlit UI
# ---------------------------------------------------------------------------

st.set_page_config(page_title="Phone Scrub", page_icon="📞", layout="centered")
st.title("Phone Scrub")
st.caption(
    "Upload a skip-trace Excel file to strip invalid, disconnected, "
    "and non-working US phone numbers."
)

# Validate API key early so the user gets a clear message
try:
    _api_key = st.secrets["IPQS_API_KEY"]
except Exception:
    st.error(
        "IPQS API key not found. "
        "Add `IPQS_API_KEY = \"your_key_here\"` to `.streamlit/secrets.toml` "
        "(local) or the app's Secrets settings on Streamlit Community Cloud."
    )
    st.stop()

uploaded_file = st.file_uploader("Upload Excel file (.xlsx)", type=["xlsx"])

col1, col2 = st.columns(2)
with col1:
    drop_voip = st.toggle(
        "Remove VOIP numbers",
        value=False,
        help="Drop numbers that IPQS identifies as VOIP",
    )
with col2:
    drop_landline = st.toggle(
        "Remove Landline numbers",
        value=False,
        help="Enable for mobile-only output",
    )

run_btn = st.button("Clean File", type="primary", disabled=uploaded_file is None)

# Persist results across reruns so the download button keeps working
if "scrub_result" not in st.session_state:
    st.session_state.scrub_result = None
if "scrub_file_id" not in st.session_state:
    st.session_state.scrub_file_id = None

# Clear stale result when a different file is uploaded
if uploaded_file:
    file_id = (uploaded_file.name, uploaded_file.size)
    if st.session_state.scrub_file_id != file_id:
        st.session_state.scrub_result = None
        st.session_state.scrub_file_id = file_id

if run_btn and uploaded_file:
    file_bytes = uploaded_file.read()
    pb = st.progress(0, text="Starting…")
    try:
        out_bytes, stats, unique_count = process_workbook(
            file_bytes, _api_key, drop_voip, drop_landline, pb
        )
        st.session_state.scrub_result = {
            "bytes": out_bytes,
            "stats": stats,
            "unique_count": unique_count,
            "filename": re.sub(r"\.xlsx$", "_cleaned.xlsx", uploaded_file.name, flags=re.I),
        }
        pb.empty()
    except ValueError as exc:
        pb.empty()
        st.error(str(exc))
    except Exception as exc:  # noqa: BLE001
        pb.empty()
        st.error(f"Unexpected error: {exc}")

if st.session_state.scrub_result:
    res = st.session_state.scrub_result
    s = res["stats"]
    removed = s["nanp"] + s["pager"] + s["inactive"] + s["type_filter"]

    st.success("File cleaned successfully.")

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Total found", s["total"])
    c2.metric("Removed", removed)
    c3.metric("Kept", s["kept"])
    c4.metric("API errors", s["api_error"], help="Kept — not dropped on lookup failure")

    with st.expander("Full breakdown"):
        st.table(
            {
                "Category": [
                    "NANP invalid",
                    "Pager",
                    "IPQS inactive / disconnected",
                    "VOIP / Landline filter",
                    "Flagged Do Not Call (kept)",
                    "Flagged leaked credentials (kept)",
                    "Unique numbers sent to IPQS",
                    "API errors (kept as Unverified)",
                ],
                "Count": [
                    s["nanp"],
                    s["pager"],
                    s["inactive"],
                    s["type_filter"],
                    s["do_not_call"],
                    s["leaked"],
                    res["unique_count"],
                    s["api_error"],
                ],
            }
        )

    st.download_button(
        label="Download Cleaned File",
        data=res["bytes"],
        file_name=res["filename"],
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
