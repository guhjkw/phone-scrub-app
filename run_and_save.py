#!/usr/bin/env python3
"""
run_and_save.py — full cleaning pipeline against the real sample file.
Saves cleaned output; prints counts and a sample of rewritten cells.
"""
from __future__ import annotations

import io
import re
import sys
import time
from collections import defaultdict

import requests
from openpyxl import load_workbook

# ── API key ──────────────────────────────────────────────────────────────────
_raw = open(".streamlit/secrets.toml").read()
API_KEY = re.search(r'IPQS_API_KEY\s*=\s*"([^"]+)"', _raw).group(1)

PHONE_RE = re.compile(r"\((\d{3})\)\s*(\d{3})-(\d{4})\s*(?:\(([A-Za-z]\w*)\))?")
N11 = {"211", "311", "411", "511", "611", "711", "811", "911"}

# ── helpers ───────────────────────────────────────────────────────────────────

def parse_cell(value):
    if value is None:
        return []
    s = str(value).strip()
    if not s or s.upper() == "N/A":
        return []
    return [
        (f"({m.group(1)}) {m.group(2)}-{m.group(3)}", m.group(4) or "")
        for m in PHONE_RE.finditer(s)
    ]

def nanp_valid(area, exchange):
    return (area[0] not in "01" and area not in N11
            and exchange[0] not in "01" and exchange not in N11)

def structural_pass(area, exchange, ptype):
    if ptype.lower() == "pager":
        return False, "pager"
    if not nanp_valid(area, exchange):
        return False, "nanp"
    return True, ""

def ipqs_lookup(digits, retries=3):
    url = f"https://ipqualityscore.com/api/json/phone/{API_KEY}/{digits}"
    for attempt in range(retries):
        try:
            resp = requests.get(url, params={"country[]": "US"}, timeout=10)
            if resp.status_code == 200:
                d = resp.json()
                if not d.get("success", False):
                    return {"error": d.get("message", "success=false")}
                return {
                    "error": None,
                    "active": d.get("active"),
                    "valid": d.get("valid"),
                    "line_type": (d.get("line_type") or "").lower(),
                    "do_not_call": bool(d.get("do_not_call", False)),
                    "leaked": bool(d.get("leaked", False)),
                }
            if resp.status_code == 429:
                time.sleep(2 ** attempt)
                continue
            return {"error": f"HTTP {resp.status_code}"}
        except Exception as exc:
            return {"error": str(exc)}
    return {"error": "max_retries"}

# ── load ──────────────────────────────────────────────────────────────────────
path = sys.argv[1] if len(sys.argv) > 1 else "sample.xlsx"
print(f"\nLoading: {path}")
wb = load_workbook(path)
ws = wb.active
print(f"Sheet: {ws.title!r}  —  {ws.max_row} rows × {ws.max_column} columns")

phone_cols = {
    cell.column
    for cell in ws[1]
    if cell.value and "phone" in str(cell.value).lower()
}
col_headers = [ws.cell(row=1, column=c).value for c in sorted(phone_cols)]
print(f"Phone columns found: {len(phone_cols)}\n")

# ── pass 1: structural filter + collect unique candidates ─────────────────────
print("Pass 1 — structural filter…")

cell_entries: dict[tuple[int,int], list[tuple[str,str]]] = {}
unique_candidates: dict[str, str] = {}  # digits → display number
s1 = defaultdict(int)

for row in ws.iter_rows(min_row=2):
    for cell in row:
        if cell.column not in phone_cols:
            continue
        parsed = parse_cell(cell.value)
        if not parsed:
            continue
        passing = []
        for number, ptype in parsed:
            s1["total"] += 1
            m = PHONE_RE.match(number)
            ok, reason = structural_pass(m.group(1), m.group(2), ptype)
            if not ok:
                s1[f"reject_{reason}"] += 1
                continue
            passing.append((number, ptype))
            digits = re.sub(r"\D", "", number)
            if digits not in unique_candidates:
                unique_candidates[digits] = number
        if passing:
            cell_entries[(cell.row, cell.column)] = passing

print(f"  Found {s1['total']} entries  →  "
      f"{s1['reject_nanp']} NANP-invalid, "
      f"{s1['reject_pager']} Pager removed  →  "
      f"{len(unique_candidates)} unique for IPQS\n")

# ── pass 2: IPQS validation ───────────────────────────────────────────────────
digits_list = list(unique_candidates.keys())
n = len(digits_list)
print(f"Pass 2 — IPQS validation ({n} unique numbers)…")

cache: dict[str, dict] = {}
errors_detail: list[str] = []

for i, digits in enumerate(digits_list, 1):
    print(f"\r  {i}/{n}  {unique_candidates[digits]:<22}", end="", flush=True)
    result = ipqs_lookup(digits)
    cache[digits] = result
    if result["error"]:
        errors_detail.append(f"    {unique_candidates[digits]} — {result['error']}")

print()  # newline

if errors_detail:
    print(f"\n  {len(errors_detail)} API error(s) — these numbers are KEPT as Unverified:")
    for e in errors_detail:
        print(e)

# ── tally per-entry stats ─────────────────────────────────────────────────────
total_kept = 0
total_removed_ipqs = 0
total_unverified = 0
dnc_entries = 0
voip_entries = 0
leaked_entries = 0
line_type_counts: dict[str, int] = defaultdict(int)

for (row_idx, col_idx), entries in cell_entries.items():
    for number, ptype in entries:
        digits = re.sub(r"\D", "", number)
        result = cache.get(digits, {"error": "missing"})
        if result["error"]:
            total_kept += 1
            total_unverified += 1
        elif result["active"] is False or result["valid"] is False:
            total_removed_ipqs += 1
        else:
            total_kept += 1
            if result["do_not_call"]:
                dnc_entries += 1
            if result["leaked"]:
                leaked_entries += 1
            lt = result["line_type"]
            if lt:
                line_type_counts[lt] += 1
            if lt == "voip":
                voip_entries += 1

# ── pass 3: rewrite cells in place ───────────────────────────────────────────
print("\nPass 3 — rewriting cells in place…")

drop_set: set[str] = set()
unverified_set: set[str] = set()
for digits, result in cache.items():
    if result["error"]:
        unverified_set.add(digits)
    elif result["active"] is False or result["valid"] is False:
        drop_set.add(digits)

rewrites_log: list[tuple[int, int, str, str]] = []  # (row, col, header, new_value)

for (row_idx, col_idx), entries in cell_entries.items():
    surviving = []
    for number, ptype in entries:
        digits = re.sub(r"\D", "", number)
        if digits in drop_set:
            continue
        label = (f"{ptype} - Unverified" if ptype else "Unverified") if digits in unverified_set else ptype
        surviving.append((number, label))

    cell = ws.cell(row=row_idx, column=col_idx)
    new_val = (
        "\n".join(f"{num} ({lbl})" if lbl else num for num, lbl in surviving)
        if surviving else None
    )
    cell.value = new_val
    hdr = ws.cell(row=1, column=col_idx).value
    rewrites_log.append((row_idx, col_idx, str(hdr), new_val or "(cleared)"))

# ── save output ───────────────────────────────────────────────────────────────
out_path = "sample_cleaned.xlsx"

# Add summary sheet
if "Summary" in wb.sheetnames:
    del wb["Summary"]
ss = wb.create_sheet("Summary")
total_removed = s1["reject_nanp"] + s1["reject_pager"] + total_removed_ipqs
for row in [
    ["Metric", "Count"],
    ["Total entries found", s1["total"]],
    ["Removed — NANP invalid", s1["reject_nanp"]],
    ["Removed — Pager", s1["reject_pager"]],
    ["Removed — IPQS inactive / disconnected", total_removed_ipqs],
    ["Total removed", total_removed],
    ["Kept (active)", total_kept],
    ["  of which: Unverified (API error, kept)", total_unverified],
    ["", ""],
    ["Flagged Do Not Call (kept, not removed)", dnc_entries],
    ["VOIP numbers (kept unless toggle on)", voip_entries],
    ["Leaked credentials (kept, not removed)", leaked_entries],
    ["", ""],
    ["Unique numbers sent to IPQS", n],
    ["IPQS API errors", len(errors_detail)],
]:
    ss.append(row)

wb.save(out_path)
print(f"  Saved to {out_path}")

# ── report ────────────────────────────────────────────────────────────────────
SEP = "─" * 56
print(f"\n{SEP}")
print("  RESULTS")
print(SEP)
print(f"  Total entries found           : {s1['total']}")
print(f"  Removed — NANP invalid        : {s1['reject_nanp']}")
print(f"  Removed — Pager               : {s1['reject_pager']}")
print(f"  Removed — IPQS inactive/disc. : {total_removed_ipqs}")
print(f"  ─────────────────────────────────────")
print(f"  Total removed                 : {total_removed}")
print(f"  Kept                          : {total_kept}")
print(f"  (of which Unverified)         : ({total_unverified})")
print(f"\n  Do Not Call flagged           : {dnc_entries}")
print(f"  VOIP                          : {voip_entries}")
print(f"  Leaked credentials            : {leaked_entries}")

if line_type_counts:
    print(f"\n  Line type breakdown (kept entries):")
    for lt, c in sorted(line_type_counts.items(), key=lambda x: -x[1]):
        print(f"    {lt:<22} {c}")

# ── sample rewritten cells ────────────────────────────────────────────────────
print(f"\n{SEP}")
print("  SAMPLE REWRITTEN PHONE CELLS  (first 8 non-empty)")
print(SEP)

shown = 0
for row_idx, col_idx, hdr, new_val in rewrites_log:
    if new_val == "(cleared)" or shown >= 8:
        continue
    print(f"\n  Row {row_idx}  [{hdr}]")
    for line in new_val.split("\n"):
        print(f"    {line}")
    shown += 1

# ── layout integrity check ────────────────────────────────────────────────────
print(f"\n{SEP}")
print("  LAYOUT CHECK")
print(SEP)
wb2 = load_workbook(out_path)
ws2 = wb2.active
print(f"  Output columns : {ws2.max_column}  (input was {ws.max_column})")
print(f"  Output rows    : {ws2.max_row}  (input was {ws.max_row})")

# spot-check: show 3 non-phone cells from row 2 to verify they're unchanged
orig_wb = load_workbook(path)
orig_ws = orig_wb.active
non_phone_cols = [
    c for c in range(1, min(20, ws.max_column + 1))
    if c not in phone_cols
][:3]
print(f"\n  Non-phone cell spot-check (row 2):")
for c in non_phone_cols:
    hdr = ws.cell(row=1, column=c).value
    orig_val = orig_ws.cell(row=2, column=c).value
    out_val  = ws2.cell(row=2, column=c).value
    match = "OK" if orig_val == out_val else "MISMATCH"
    print(f"    col {c:3d} [{hdr}]:  {str(orig_val)[:40]!r}  →  {match}")

print(f"\n{SEP}\n")
