#!/usr/bin/env python3
"""
run_real.py — run the phone-scrub pipeline against the real sample file.
Reports totals only; removal rule unchanged (drop if active=False OR valid=False).
"""
from __future__ import annotations

import io
import re
import sys
import time
from collections import defaultdict

import requests
from openpyxl import load_workbook

# ── API key ──────────────────────────────────────────────────────────────────
_raw = open(".streamlit/secrets.toml").read()
API_KEY = re.search(r'IPQS_API_KEY\s*=\s*"([^"]+)"', _raw).group(1)

PHONE_RE = re.compile(r"\((\d{3})\)\s*(\d{3})-(\d{4})\s*(?:\(([A-Za-z]\w*)\))?")
N11 = {"211", "311", "411", "511", "611", "711", "811", "911"}

# ── helpers (same as app.py) ─────────────────────────────────────────────────

def parse_cell(value):
    if value is None:
        return []
    s = str(value).strip()
    if not s or s.upper() == "N/A":
        return []
    return [
        (f"({m.group(1)}) {m.group(2)}-{m.group(3)}", m.group(4) or "")
        for m in PHONE_RE.finditer(s)
    ]

def nanp_valid(area, exchange):
    return (area[0] not in "01" and area not in N11
            and exchange[0] not in "01" and exchange not in N11)

def structural_pass(area, exchange, ptype):
    if ptype.lower() == "pager":
        return False, "pager"
    if not nanp_valid(area, exchange):
        return False, "nanp"
    return True, ""

def ipqs_lookup(digits, retries=3):
    url = f"https://ipqualityscore.com/api/json/phone/{API_KEY}/{digits}"
    for attempt in range(retries):
        try:
            resp = requests.get(url, params={"country[]": "US"}, timeout=10)
            if resp.status_code == 200:
                d = resp.json()
                if not d.get("success", False):
                    return {"error": d.get("message", "success=false")}
                return {
                    "error": None,
                    "active": d.get("active"),
                    "valid": d.get("valid"),
                    "line_type": (d.get("line_type") or "").lower(),
                    "do_not_call": bool(d.get("do_not_call", False)),
                    "leaked": bool(d.get("leaked", False)),
                }
            if resp.status_code == 429:
                time.sleep(2 ** attempt)
                continue
            return {"error": f"HTTP {resp.status_code}"}
        except Exception as exc:
            return {"error": str(exc)}
    return {"error": "max_retries"}

# ── load workbook ─────────────────────────────────────────────────────────────
path = sys.argv[1] if len(sys.argv) > 1 else "sample.xlsx"
print(f"\nLoading: {path}")
wb = load_workbook(path)
ws = wb.active
print(f"Sheet: {ws.title!r}  —  {ws.max_row} rows × {ws.max_column} columns")

phone_cols = {
    cell.column
    for cell in ws[1]
    if cell.value and "phone" in str(cell.value).lower()
}
col_headers = [ws.cell(row=1, column=c).value for c in sorted(phone_cols)]
print(f"Phone columns ({len(phone_cols)}): {col_headers}\n")

# ── pass 1: structural filter ─────────────────────────────────────────────────
stats = defaultdict(int)
unique_candidates: dict[str, str] = {}   # digits → display number

for row in ws.iter_rows(min_row=2):
    for cell in row:
        if cell.column not in phone_cols:
            continue
        for number, ptype in parse_cell(cell.value):
            stats["total_found"] += 1
            m = PHONE_RE.match(number)
            ok, reason = structural_pass(m.group(1), m.group(2), ptype)
            if not ok:
                stats[f"reject_{reason}"] += 1
                continue
            stats["pass_structural"] += 1
            digits = re.sub(r"\D", "", number)
            if digits not in unique_candidates:
                unique_candidates[digits] = number

print(f"Pass 1 — structural filter")
print(f"  Total entries found    : {stats['total_found']}")
print(f"  Rejected NANP invalid  : {stats['reject_nanp']}")
print(f"  Rejected Pager         : {stats['reject_pager']}")
print(f"  Pass structural        : {stats['pass_structural']}")
print(f"  Unique → IPQS          : {len(unique_candidates)}\n")

# ── pass 2: IPQS validation ───────────────────────────────────────────────────
cache: dict[str, dict] = {}
n = len(unique_candidates)
digits_list = list(unique_candidates.keys())

print(f"Pass 2 — IPQS validation ({n} unique numbers)…")
for i, digits in enumerate(digits_list, 1):
    pct = i / n
    bar = "█" * int(pct * 40)
    print(f"\r  [{bar:<40}] {i}/{n}", end="", flush=True)
    cache[digits] = ipqs_lookup(digits)
print()  # newline after progress bar

# ── tally results ─────────────────────────────────────────────────────────────
removed_inactive = 0
kept = 0
api_errors = 0
dnc = 0
leaked = 0
voip = 0

# Track line_type distribution
line_types: dict[str, int] = defaultdict(int)

for digits, result in cache.items():
    if result["error"]:
        api_errors += 1
        kept += 1   # kept as Unverified on error
        continue

    lt = result["line_type"]
    line_types[lt or "unknown"] += 1

    if result["do_not_call"]:
        dnc += 1
    if result["leaked"]:
        leaked += 1
    if lt == "voip":
        voip += 1

    if result["active"] is False or result["valid"] is False:
        removed_inactive += 1
    else:
        kept += 1

# Numbers removed from total_found:
# structural rejects are already in stats; removed_inactive is from IPQS pass
total_removed = stats["reject_nanp"] + stats["reject_pager"] + removed_inactive

SEP = "─" * 52
print(f"\n{SEP}")
print("  RESULTS")
print(SEP)
print(f"  Total entries found           : {stats['total_found']}")
print(f"  Removed — NANP invalid        : {stats['reject_nanp']}")
print(f"  Removed — Pager               : {stats['reject_pager']}")
print(f"  Removed — IPQS inactive/disc. : {removed_inactive}")
print(f"  ─────────────────────────────")
print(f"  Total removed                 : {total_removed}")
print(f"  Kept (active numbers)         : {kept}")
print(f"  API errors (kept, unverified) : {api_errors}")
print(f"\n  Of the {n} unique numbers validated:")
print(f"  Do Not Call flagged           : {dnc}")
print(f"  VOIP                          : {voip}")
print(f"  Leaked credentials            : {leaked}")

if line_types:
    print(f"\n  Line type breakdown:")
    for lt, count in sorted(line_types.items(), key=lambda x: -x[1]):
        print(f"    {lt:<20} {count}")

print(f"{SEP}\n")
