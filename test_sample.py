#!/usr/bin/env python3
"""
test_sample.py — end-to-end smoke test with a synthetic 10-row sample.
Runs without Streamlit.  Uses the real IPQS API key from secrets.toml.
"""
from __future__ import annotations

import io
import re
import time
import pprint

import requests
from openpyxl import Workbook, load_workbook

# ---------------------------------------------------------------------------
# Read API key from local secrets.toml (gitignored)
# ---------------------------------------------------------------------------
_raw = open(".streamlit/secrets.toml").read()
API_KEY = re.search(r'IPQS_API_KEY\s*=\s*"([^"]+)"', _raw).group(1)

# ---------------------------------------------------------------------------
# Core helpers (mirrors app.py — kept in sync manually)
# ---------------------------------------------------------------------------
PHONE_RE = re.compile(r"\((\d{3})\)\s*(\d{3})-(\d{4})\s*(?:\(([A-Za-z]\w*)\))?")
N11 = {"211", "311", "411", "511", "611", "711", "811", "911"}


def parse_cell(value):
    if value is None:
        return []
    s = str(value).strip()
    if not s or s.upper() == "N/A":
        return []
    return [
        (f"({m.group(1)}) {m.group(2)}-{m.group(3)}", m.group(4) or "")
        for m in PHONE_RE.finditer(s)
    ]


def nanp_valid(area, exchange):
    return (
        area[0] not in "01"
        and area not in N11
        and exchange[0] not in "01"
        and exchange not in N11
    )


def structural_pass(area, exchange, ptype):
    if ptype.lower() == "pager":
        return False, "pager"
    if not nanp_valid(area, exchange):
        return False, "nanp"
    return True, ""


def ipqs_lookup(digits, retries=3):
    url = f"https://ipqualityscore.com/api/json/phone/{API_KEY}/{digits}"
    for attempt in range(retries):
        try:
            resp = requests.get(url, params={"country[]": "US"}, timeout=10)
            if resp.status_code == 200:
                d = resp.json()
                if not d.get("success", False):
                    return {"error": d.get("message", "success=false"), "_raw": d}
                return {
                    "error": None,
                    "_raw": d,
                    "active": d.get("active"),
                    "valid": d.get("valid"),
                    "line_type": (d.get("line_type") or "").lower(),
                    "do_not_call": bool(d.get("do_not_call", False)),
                    "leaked": bool(d.get("leaked", False)),
                }
            if resp.status_code == 429:
                time.sleep(2 ** attempt)
                continue
            return {"error": f"HTTP {resp.status_code}", "_raw": {}}
        except Exception as exc:  # noqa: BLE001
            return {"error": str(exc), "_raw": {}}
    return {"error": "max_retries", "_raw": {}}


# ---------------------------------------------------------------------------
# Build the synthetic 10-row sample workbook
# ---------------------------------------------------------------------------

def build_sample() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(["First Name", "Last Name", "Street", "Phone 1", "Phone 2", "Email"])
    rows = [
        # two numbers in one cell — tests multi-parse and dedup
        ["Alice", "Smith", "123 Main St",
         "(661) 600-2347 (Wireless)\n\n(661) 297-0544 (Landline)", "", "alice@x.com"],
        # two columns with different numbers
        ["Bob", "Jones", "456 Oak Ave",
         "(213) 621-0002 (Wireless)", "(310) 285-2500 (Landline)", "bob@x.com"],
        # NANP-invalid area code (starts with 1) — structural filter stops this
        ["Carol", "White", "789 Pine Rd",
         "(100) 555-1234 (Wireless)", "", "carol@x.com"],
        # Pager — structural filter stops this
        ["Dave", "Brown", "321 Elm St",
         "(818) 999-1234 (Pager)", "", "dave@x.com"],
        # N/A cell
        ["Eve", "Davis", "654 Maple Dr", "N/A", "", "eve@x.com"],
        # blank cell
        ["Frank", "Miller", "987 Cedar Ln", "", "", "frank@x.com"],
        # exact duplicate of Alice row 1 phone — must NOT trigger a second IPQS call
        ["Grace", "Wilson", "147 Birch Blvd",
         "(661) 600-2347 (Wireless)", "", "grace@x.com"],
        # second-column duplicate of Bob's first number
        ["Hank", "Moore", "258 Spruce Way",
         "(415) 555-2671 (Wireless)", "(213) 621-0002 (Landline)", "hank@x.com"],
        # no type label at all
        ["Iris", "Taylor", "369 Ash Ave", "(323) 456-7890", "", "iris@x.com"],
        # N11 area code (211) — structural filter
        ["Jack", "Anderson", "741 Walnut Rd",
         "(211) 555-1234 (Wireless)", "", "jack@x.com"],
    ]
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Run the test
# ---------------------------------------------------------------------------

SEP = "─" * 62

print(f"\n{SEP}")
print("  PHONE SCRUB  —  sample test")
print(SEP)

sample_bytes = build_sample()

# ── Step 1: parse + structural filter ───────────────────────────────────────
print("\nStep 1 · Parsing & structural filter\n")

wb = load_workbook(io.BytesIO(sample_bytes))
ws = wb.active

phone_cols = {
    cell.column
    for cell in ws[1]
    if cell.value and "phone" in str(cell.value).lower()
}
col_names = [ws.cell(row=1, column=c).value for c in sorted(phone_cols)]
print(f"  Phone columns detected: {col_names}  (col indices {sorted(phone_cols)})\n")

unique_candidates: dict[str, str] = {}   # digits → display number
structural_log: list[str] = []

for row in ws.iter_rows(min_row=2):
    for cell in row:
        if cell.column not in phone_cols:
            continue
        for number, ptype in parse_cell(cell.value):
            m = PHONE_RE.match(number)
            area, exchange = m.group(1), m.group(2)
            ok, reason = structural_pass(area, exchange, ptype)
            label = ptype or "(no type)"
            if ok:
                digits = re.sub(r"\D", "", number)
                is_dup = digits in unique_candidates
                structural_log.append(
                    f"  row {cell.row:2d}  {number} ({label})  →  PASS"
                    + ("  [dup, no extra API call]" if is_dup else "")
                )
                if not is_dup:
                    unique_candidates[digits] = number
            else:
                structural_log.append(
                    f"  row {cell.row:2d}  {number} ({label})  →  REJECT ({reason})"
                )

for line in structural_log:
    print(line)

print(f"\n  Unique candidates for IPQS: {len(unique_candidates)}")

# ── Step 2: IPQS validation ──────────────────────────────────────────────────
print(f"\n{SEP}")
print("Step 2 · IPQS validation  (one call per unique number)\n")

cache: dict[str, dict] = {}
api_call_count = 0

for digits, number in unique_candidates.items():
    print(f"  → {number}  ({digits})")
    result = ipqs_lookup(digits)
    cache[digits] = result
    api_call_count += 1
    if result["error"]:
        print(f"       ERROR: {result['error']}")
    else:
        r = result["_raw"]
        print(
            f"       valid={result['valid']}  active={result['active']}  "
            f"line_type={result['line_type']!r}"
        )
        print(
            f"       do_not_call={result['do_not_call']}  "
            f"leaked={result['leaked']}  "
            f"fraud_score={r.get('fraud_score')}  "
            f"carrier={r.get('carrier')!r}"
        )

print(f"\n  API calls made: {api_call_count}  (expected ≤ {len(unique_candidates)})")

# ── Step 3: apply removal rules + rebuild cells ──────────────────────────────
print(f"\n{SEP}")
print("Step 3 · Apply rules (remove if active=False OR valid=False)\n")

# Reload so we can rewrite
wb2 = load_workbook(io.BytesIO(sample_bytes))
ws2 = wb2.active

drop_map: dict[str, str] = {}
unverified: set[str] = set()

for digits, result in cache.items():
    if result["error"]:
        unverified.add(digits)
        continue
    if result["active"] is False or result["valid"] is False:
        drop_map[digits] = "inactive"

cell_results: list[str] = []
stats = dict(total=0, kept=0, removed=0, unverified=0,
             do_not_call=0, leaked=0, api_error=0,
             nanp=sum(1 for l in structural_log if "REJECT (nanp)" in l),
             pager=sum(1 for l in structural_log if "REJECT (pager)" in l))

for digits, result in cache.items():
    if not result["error"]:
        if result["do_not_call"]:
            stats["do_not_call"] += 1
        if result["leaked"]:
            stats["leaked"] += 1
    else:
        stats["api_error"] += 1

cell_entries: dict[tuple[int, int], list[tuple[str, str]]] = {}
for row in ws2.iter_rows(min_row=2):
    for cell in row:
        if cell.column not in phone_cols:
            continue
        parsed = parse_cell(cell.value)
        if not parsed:
            continue
        passing = []
        for number, ptype in parsed:
            m = PHONE_RE.match(number)
            ok, _ = structural_pass(m.group(1), m.group(2), ptype)
            if ok:
                passing.append((number, ptype))
                stats["total"] += 1
        if passing:
            cell_entries[(cell.row, cell.column)] = passing

for (row_idx, col_idx), entries in cell_entries.items():
    surviving = []
    for number, ptype in entries:
        digits = re.sub(r"\D", "", number)
        if digits in drop_map:
            stats["removed"] += 1
            cell_results.append(
                f"  row {row_idx:2d} col {col_idx}  "
                f"{number} ({ptype or 'no-type'})  →  REMOVED (inactive)"
            )
            continue
        if digits in unverified:
            label = f"{ptype} - Unverified" if ptype else "Unverified"
            stats["unverified"] += 1
        else:
            label = ptype
        surviving.append((number, label))
        stats["kept"] += 1
        cell_results.append(
            f"  row {row_idx:2d} col {col_idx}  "
            f"{number} ({ptype or 'no-type'})  →  KEPT  "
            + (f"[label: {label!r}]" if label != ptype else "")
        )
    cell = ws2.cell(row=row_idx, column=col_idx)
    cell.value = (
        "\n".join(
            f"{n} ({lbl})" if lbl else n
            for n, lbl in surviving
        ) if surviving else None
    )

for line in sorted(cell_results):
    print(line)

# ── Summary ──────────────────────────────────────────────────────────────────
print(f"\n{SEP}")
print("Summary\n")

rows_summary = [
    ("Total phone entries (post-structural)", stats["total"]),
    ("Removed — NANP invalid",               stats["nanp"]),
    ("Removed — Pager",                       stats["pager"]),
    ("Removed — IPQS inactive/disconnected",  stats["removed"]),
    ("Kept",                                  stats["kept"]),
    ("  of which: Unverified (API error)",    stats["unverified"]),
    ("Flagged Do Not Call  (not removed)",    stats["do_not_call"]),
    ("Flagged leaked creds (not removed)",    stats["leaked"]),
    ("IPQS API errors",                       stats["api_error"]),
    ("Unique numbers sent to IPQS",           api_call_count),
]
w = max(len(r[0]) for r in rows_summary)
for label, count in rows_summary:
    print(f"  {label:<{w}}  {count}")

# ── Save output for inspection ────────────────────────────────────────────────
out_path = "test_output.xlsx"
out = io.BytesIO()
wb2.save(out)
with open(out_path, "wb") as f:
    f.write(out.getvalue())
print(f"\n  Output saved to {out_path}")
print(f"\n{SEP}\n")
